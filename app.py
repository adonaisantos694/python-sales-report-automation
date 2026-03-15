from pathlib import Path
from typing import List, Dict, Any
import csv
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.chart import BarChart, Reference


PRODUCT_COLUMN_WIDTH: int = 35
INSIGHT_COLUMN_WIDTH: int = 80


def parse_price(value: Any) -> float:
    if value is None:
        return 0.0

    cleaned = re.sub(r"[^\d.]", "", str(value))

    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_int(value: Any) -> int:
    if value is None:
        return 0

    cleaned = re.sub(r"[^\d]", "", str(value))

    try:
        return int(cleaned)
    except ValueError:
        return 0


def parse_float(value: Any) -> float:
    if value is None:
        return 0.0

    cleaned = re.sub(r"[^\d.]", "", str(value))

    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def load_csv_data(file_path: Path) -> List[Dict[str, Any]]:
    data: List[Dict[str, Any]] = []

    with file_path.open("r", encoding="utf-8") as file:
        reader = csv.DictReader(file)

        for row in reader:
            data.append(row)

    return data


def simplify_dataset(data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    simplified: List[Dict[str, Any]] = []

    for row in data:
        product = str(row.get("product_name", ""))[:100]
        category = str(row.get("category", "")).split("|")[0]

        discount_price = parse_price(row.get("discounted_price"))
        original_price = parse_price(row.get("actual_price"))

        rating = parse_float(row.get("rating"))
        rating_count = parse_int(row.get("rating_count"))

        revenue_estimate = discount_price * rating_count
        discount_value = original_price - discount_price

        simplified.append(
            {
                "product": product,
                "category": category,
                "discount_price": discount_price,
                "original_price": original_price,
                "discount_value": discount_value,
                "rating": rating,
                "rating_count": rating_count,
                "revenue_estimate": revenue_estimate,
            }
        )

    return simplified


def analyze_data(data: List[Dict[str, Any]]) -> Dict[str, Any]:

    avg_rating = sum(d["rating"] for d in data) / len(data)

    most_reviewed = max(data, key=lambda x: x["rating_count"])
    best_rated = max(data, key=lambda x: x["rating"])
    biggest_discount = max(data, key=lambda x: x["discount_value"])
    highest_revenue = max(data, key=lambda x: x["revenue_estimate"])

    category_count: Dict[str, int] = {}
    category_rating: Dict[str, List[float]] = {}

    for row in data:
        category = row["category"]

        category_count[category] = category_count.get(category, 0) + 1
        category_rating.setdefault(category, []).append(row["rating"])

    top_category = max(category_count, key=category_count.get)

    best_category = max(
        category_rating,
        key=lambda c: sum(category_rating[c]) / len(category_rating[c]),
    )

    return {
        "average_rating": round(avg_rating, 2),
        "most_reviewed_product": most_reviewed["product"],
        "best_rated_product": best_rated["product"],
        "biggest_discount_product": biggest_discount["product"],
        "highest_revenue_product": highest_revenue["product"],
        "category_with_most_products": top_category,
        "best_rated_category": best_category,
        "category_counts": category_count,
    }


def format_sheet(sheet, width: int) -> None:
    for column in sheet.columns:
        letter = column[0].column_letter

        for cell in column:
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        sheet.column_dimensions[letter].width = width


def create_charts(workbook, category_counts):

    chart_sheet = workbook.create_sheet("CHART_DATA")

    chart_sheet.append(["Category", "Product Count"])

    for category, count in category_counts.items():
        chart_sheet.append([category, count])

    chart = BarChart()
    chart.title = "Products per Category"
    chart.y_axis.title = "Products"
    chart.x_axis.title = "Category"

    data = Reference(chart_sheet, min_col=2, min_row=1,
                     max_row=len(category_counts) + 1)
    categories = Reference(chart_sheet, min_col=1,
                           min_row=2, max_row=len(category_counts) + 1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)

    chart_sheet.add_chart(chart, "E2")


def create_excel_report(
    data: List[Dict[str, Any]],
    insights: Dict[str, Any],
    output_file: Path,
) -> None:

    workbook = Workbook()

    products_sheet = workbook.active
    products_sheet.title = "PRODUCTS"

    headers = list(data[0].keys())
    products_sheet.append(headers)

    for row in data:
        products_sheet.append([row.get(h) for h in headers])

    format_sheet(products_sheet, PRODUCT_COLUMN_WIDTH)

    insight_sheet = workbook.create_sheet("INSIGHTS")

    insight_sheet.append(["BUSINESS QUESTION", "INSIGHT"])

    insight_sheet.append(
        [
            "What is the average product rating across the dataset?",
            insights["average_rating"],
        ]
    )

    insight_sheet.append(
        [
            "Which product has the highest number of customer reviews?",
            insights["most_reviewed_product"],
        ]
    )

    insight_sheet.append(
        [
            "Which product has the highest rating?",
            insights["best_rated_product"],
        ]
    )

    insight_sheet.append(
        [
            "Which product offers the largest discount?",
            insights["biggest_discount_product"],
        ]
    )

    insight_sheet.append(
        [
            "Which product generates the highest estimated revenue?",
            insights["highest_revenue_product"],
        ]
    )

    insight_sheet.append(
        [
            "Which category contains the largest number of products?",
            insights["category_with_most_products"],
        ]
    )

    insight_sheet.append(
        [
            "Which category has the highest average rating?",
            insights["best_rated_category"],
        ]
    )

    format_sheet(insight_sheet, INSIGHT_COLUMN_WIDTH)

    create_charts(workbook, insights["category_counts"])

    workbook.save(output_file)


def main() -> None:

    input_file = Path("amazon.csv")
    output_file = Path("final_report.xlsx")

    if not input_file.exists():
        raise FileNotFoundError(f"File not found: {input_file}")

    raw_data = load_csv_data(input_file)

    simplified_data = simplify_dataset(raw_data)

    insights = analyze_data(simplified_data)

    create_excel_report(
        simplified_data,
        insights,
        output_file,
    )


if __name__ == "__main__":
    main()
